import json
# Code to selectively color some lines


fileOpen1 = open("myJson.json", "r")
details = json.loads(fileOpen1.read())

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

print("Reading .....")
wb = load_workbook(filename = 'ExcelWork.xlsx')
ws = wb.active
print("Starting ...")

redFill = PatternFill(start_color='00FF0000',
                       end_color='00FF0000',
                       fill_type='solid'
                      )

coloured = 0
for i in range(0, len(details)):
    if coloured%50 == 0:
        print("Coloured = " + str(coloured))

    if details[i][1] == False:
        if details[i][0] == ws["K" + str(i + 4)].value:
            coloured += 1
            for j in range(0, 26):
                ws[chr(j + 65) + str(i + 4)].fill = redFill
                ws['Z' + str(i + 4)].value = "Mail Not Found "

print(str(coloured))

wb.save('ExcelWork1.xlsx')
