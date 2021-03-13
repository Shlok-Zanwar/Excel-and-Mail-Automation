# To automate checking of Excel and color them based of conditions

import json
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

fileOpen1 = open("itemCodes2.json", "r")
details = json.loads(fileOpen1.read())
print(len(details))

wb = load_workbook(filename = 'ExcelWork.xlsx')
ws = wb.active


print("Starting ... ")



store = 4

orangeFill = PatternFill(start_color='00FF9900',
                       end_color='00FF9900',
                       fill_type='solid'
                      )


pinkFill = PatternFill(start_color='00FF99CC',
                       end_color='00FF99CC',
                       fill_type='solid'
                      )



for detail in details:
    if detail[1] == 1:
        for i in range(store, 7194):
            if detail[0] == ws["F"+str(i)].value:
                inQty = float(ws["M"+str(i)].value)
                outQty = float(ws["O"+str(i)].value)

                if inQty <= 1000:

                    if outQty > inQty*1.10:
                        for j in range(0, 26):
                            ws[chr(j + 65) + str(i)].fill = orangeFill

                        ws['Z' + str(i)].value = "Excess Quantity "
                        store = i
                        break

                    else:
                        for j in range(0, 26):
                            ws[chr(j + 65) + str(i)].fill = pinkFill

                        store = i
                        break

                elif inQty > 1000 and inQty <= 10000:

                    if outQty > inQty * 1.05:
                        for j in range(0, 26):
                            ws[chr(j + 65) + str(i)].fill = orangeFill

                        ws['Z' + str(i)].value = "Excess Quantity "
                        store = i
                        break

                    else:
                        for j in range(0, 26):
                            ws[chr(j + 65) + str(i)].fill = pinkFill

                        store = i
                        break

                elif inQty > 10000:

                    if outQty > inQty * 1.025:
                        for j in range(0, 26):
                            ws[chr(j + 65) + str(i)].fill = orangeFill

                        ws['Z' + str(i)].value = "Excess Quantity "
                        store = i
                        break

                    else:
                        for j in range(0, 26):
                            ws[chr(j + 65) + str(i)].fill = pinkFill

                        store = i
                        break

print("Done")
wb.save('ExcelWork4.xlsx')
